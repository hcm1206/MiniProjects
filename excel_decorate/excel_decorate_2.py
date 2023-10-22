# 액셀 파일의 셀 값에 따라 배경색 추가

import openpyxl
from openpyxl.styles import PatternFill

# 액셀 파일의 데이터를 로드
wb = openpyxl.load_workbook('./test.xlsx', data_only=True)
# 액셀 파일에서 Sheet1 시트를 선택
ws = wb["Sheet1"]

# RGB 헥사 코드가 99ff99(초록색)인 good_color 색 정보 저장
good_color = PatternFill(start_color='99ff99', end_color='99ff99', fill_type='solid')
# RGB 헥사 코드가 ffff99(노란색)인 normal_color 색 정보 저장
normal_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
# RGB 헥사 코드가 ff9999(빨간색)인 bad_color 색 정보 저장
bad_color = PatternFill(start_color='ff9999', end_color='ff9999', fill_type='solid')

# 액셀 파일 Sheet1 시트의 C열 전체를 선택
cells = ws['C':'C']

# C열의 각 셀마다 반복
for cell in cells:
    # 셀에 값이 존재하면
    try:
        # 셀의 숫자 값을 받아옴
        value = int(cell.value)
        # 셀의 숫자 값이 10000 이상이면
        if value >= 10000:
            # 해당 셀의 배경색을 good_color(초록색)로 설정
            cell.fill = good_color
        # 셀의 숫자 값이 1000~10000이면
        elif value >= 1000:
            # 해당 셀의 배경색을 normal_color(노란색)로 설정
            cell.fill = normal_color
        # 셀의 숫자 값이 1000 미만이면
        else:
            # 해당 셀의 배경색을 bad_color(빨간색)로 설정
            cell.fill = bad_color
    # 셀에 값이 없으면 무시
    except:
        pass

# 설정한 내용의 액셀 정보를 xlsx 파일로 별도 저장
wb.save('test_result.xlsx')