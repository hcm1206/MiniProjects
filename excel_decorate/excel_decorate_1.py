# 액셀 파일의 특정 셀에 배경색 추가

import openpyxl
from openpyxl.styles import PatternFill

# 액셀 파일의 데이터를 로드
wb = openpyxl.load_workbook('./test.xlsx', data_only=True)
# 액셀 파일에서 Sheet1 시트를 선택
ws = wb["Sheet1"]
# RGB 헥사 코드가 99FF99인 색 정보 color 설정
color = PatternFill(start_color='99ff99', end_color='99ff99', fill_type='solid')
# Sheet1 시트에서 3행 3열에 해당하는 셀에 설정한 color 색 정보을 적용
ws.cell(3,3).fill = color
# 설정한 내용을 xlsx 파일로 별도 저장
wb.save('test_result.xlsx')